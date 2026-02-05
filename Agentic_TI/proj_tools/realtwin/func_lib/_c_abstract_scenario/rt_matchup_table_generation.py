##############################################################################
# Copyright (c) 2024, Oak Ridge National Laboratory                          #
# All rights reserved.                                                       #
#                                                                            #
# This file is part of RealTwin and is distributed under a GPL               #
# license. For the licensing terms see the LICENSE file in the top-level     #
# directory.                                                                 #
#                                                                            #
# Contributors: ORNL Real-Twin Team                                          #
# Contact: realtwin@ornl.gov                                                 #
##############################################################################

import xml.etree.ElementTree as ET
import pandas as pd
from pyproj import Proj, Transformer
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment
import copy


def get_net_edges(path_net: str) -> pd.DataFrame:
    """ Extract edges from a SUMO network XML file and return them as a DataFrame."""
    # Load the network XML file
    tree = ET.parse(path_net)
    root = tree.getroot()

    # Extract edges and store them in a DataFrame
    edges_data = []
    for edge in root.findall("edge"):
        edge_id = edge.get("id")
        from_junction = edge.get("from")
        to_junction = edge.get("to")
        edges_data.append(
            {"Edge ID": edge_id, "From": from_junction, "To": to_junction})
    df_edges = pd.DataFrame(edges_data)
    return df_edges


def get_net_connections(path_net: str) -> pd.DataFrame:
    """ Extract connections from a SUMO network XML file and return them as a DataFrame."""
    # Load the network XML file
    tree = ET.parse(path_net)
    root = tree.getroot()
    connections_data = []
    direction_mapping = {"s": "thru",
                         "l": "left", "L": "left",
                         "r": "right", "R": "right",
                         "t": "Uturn", "invalid": "invalid"}
    for connection in root.findall("connection"):
        from_edge = connection.get("from")
        to_edge = connection.get("to")
        direction_code = connection.get("dir")
        direction = direction_mapping.get(direction_code, None)
        connections_data.append({"FromEdge": from_edge, "ToEdge": to_edge, "Direction": direction})

    df_connections = pd.DataFrame(connections_data)
    return df_connections


def generate_matchup_table(df_matchup_table: pd.DataFrame, path_output: str = "MatchUp_Table.xlsx") -> bool:
    """ Generate a matchup table from the provided DataFrame and save it to an Excel file."""
    network_columns = ["JunctionID_OpenDrive", "Bearing", "Numbering", "FromRoadID_OpenDrive", "ToRoadID_OpenDrive", "Turn"]
    demand_columns = ["File_GridSmart", "Date_GridSmart", "IntersectionName_GridSmart", "Turn_GridSmart"]
    signal_columns = ["File_Synchro", "IntersectionID_Synchro", "Turn_Synchro"]
    other_columns = ["Need calibration?"]

    # df_matchup_table = path_net
    # Generate the matchup table
    # df_matchup_table = format_junction_bearing(path_net)

    wb = Workbook()
    ws = wb.active

    ws.append(["Network"] * len(network_columns) + ["Demand"] * len(demand_columns) +
              ["Signal"] * len(signal_columns) + [""] * len(other_columns))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(network_columns))
    ws.merge_cells(start_row=1, start_column=len(network_columns) + 1, end_row=1,
                   end_column=len(network_columns) + len(demand_columns))
    ws.merge_cells(start_row=1, start_column=len(network_columns) + len(demand_columns) + 1, end_row=1,
                   end_column=len(network_columns) + len(demand_columns) + len(signal_columns))

    ws.append(network_columns + demand_columns + signal_columns + other_columns)

    for row in df_matchup_table.itertuples(index=False):
        ws.append(list(row))

    current_start = 3  # Data starts at row 3
    for i in range(3, len(df_matchup_table) + 3):
        if (i == len(df_matchup_table) + 2 or ws[f"A{i}"].value != ws[f"A{i + 1}"].value):  # Check next row
            if current_start < i:  # Only merge if there are multiple same values
                ws.merge_cells(start_row=current_start, start_column=1, end_row=i, end_column=1)  # JunctionID_OpenDrive
                ws.merge_cells(start_row=current_start, start_column=7, end_row=i, end_column=7)  # File_GridSmart
                ws.merge_cells(start_row=current_start, start_column=8, end_row=i, end_column=8)  # Date_GridSmart
                ws.merge_cells(start_row=current_start, start_column=9, end_row=i, end_column=9)  # IntersectionName_GridSmart
                # ws.merge_cells(start_row=current_start, start_column=11, end_row=i, end_column=11)  # File_Synchro
                ws.merge_cells(start_row=current_start, start_column=12, end_row=i, end_column=12)  # IntersectionID_Synchro
                ws.merge_cells(start_row=current_start, start_column=14, end_row=i, end_column=14)  # Need calibration?
            current_start = i + 1

    if len(df_matchup_table) > 0:
        ws.merge_cells(start_row=3, start_column=11, end_row=len(
            df_matchup_table) + 2, end_column=11)

    # Center align merged cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    column_widths = {
        "A": 20,  # JunctionID_OpenDrive
        "B": 15,  # Bearing
        "C": 15,  # Numbering
        "D": 25,  # FromRoadID_OpenDrive
        "E": 25,  # ToRoadID_OpenDrive
        "F": 15,  # Turn
        "G": 20,  # File_GridSmart
        "H": 20,  # Date_GridSmart
        "I": 30,  # IntersectionName_GridSmart
        "J": 20,  # Turn_GridSmart
        "K": 20,  # File_Synchro
        "L": 25,  # IntersectionID_Synchro
        "M": 20,  # Turn_Synchro
        "N": 20   # Need calibration?
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    wb.save(path_output)
    return True


def generate_junction_bearing(path_net: str) -> pd.DataFrame:
    """ Generate a DataFrame containing junction bearings from a SUMO network XML file."""
    tree = ET.parse(path_net)
    root = tree.getroot()

    # Extract edges and store them in a DataFrame
    df_edges = get_net_edges(path_net)

    # Extract correct UTM zone and netOffset
    location = root.find("location")
    if location is not None:
        net_offset_x, net_offset_y = [float(val) for val in location.get("netOffset", "0,0").split(",")]

        # Extract the UTM zone from the projection parameters
        proj_params = location.get("projParameter", "")
        utm_zone = None
        for param in proj_params.split():
            if "+zone=" in param:
                utm_zone = int(param.split("=")[1])
                break

    if utm_zone is None:
        raise ValueError("UTM zone could not be found in the XML file's projection parameters.")

    # Define the correct UTM projection with the extracted zone
    utm_proj = Proj(proj="utm", zone=utm_zone, ellps="WGS84", datum="WGS84", units="m", no_defs=True)
    wgs84_proj = Proj(proj="latlong", datum="WGS84")

    # Function to convert (x, y) to (lat, lon) using extracted net_offset and UTM zone
    def convert_coordinates(x, y):
        utm_x = x - net_offset_x
        utm_y = y - net_offset_y
        lon, lat = Transformer.from_proj(utm_proj, wgs84_proj, always_xy=True).transform(utm_x, utm_y)
        return lat, lon

    # Dictionary to store junctions with at least 3 approaches and their bearing coordinates
    junction_bearing = []

    for junction in root.findall("junction"):
        junction_id = junction.get("id")
        inc_lanes = junction.get("incLanes", "").split()

        # Extract unique edges from lane IDs
        edges = set(lane.rsplit("_", 1)[0] for lane in inc_lanes if lane)

        # Get entrance and exit counts for this junction
        entrance_count = sum(df_edges["To"] == junction_id)
        exit_count = sum(df_edges["From"] == junction_id)

        # Check if the junction has at least 2 entrances or at least 2 exits
        if entrance_count >= 2 or exit_count >= 2:
            for edge in root.findall("edge"):
                edge_id = edge.get("id")
                if edge_id in edges:
                    # Get the first lane of this edge
                    first_lane = edge.find("lane")
                    if first_lane is not None:
                        shape = first_lane.get("shape", "").split()

                        # Extract the last two points
                        if len(shape) >= 2:
                            last_point = shape[-1].split(",")[:2]  # Get x, y of the last point
                            second_last_point = shape[-2].split(",")[:2]  # Get x, y of the second last point

                            second_last_point = tuple([float(val) for val in second_last_point])
                            last_point = tuple([float(val) for val in last_point])

                            # Convert to lat, lon
                            lat1, lon1 = convert_coordinates(*second_last_point)
                            lat2, lon2 = convert_coordinates(*last_point)

                            # Calculate bearing
                            bearing = calculate_bearing(lat1, lon1, lat2, lon2)

                            # Store the approach with its coordinates
                            junction_bearing.append({
                                "Junction ID": junction_id,
                                "Approach Edge": edge_id,
                                "Second Last Point (x, y)": second_last_point,
                                "Last Point (x, y)": last_point,
                                "Second Last Point (lat, lon)": (lat1, lon1),
                                "Last Point (lat, lon)": (lat2, lon2),
                                "Degree": round(bearing, 2),
                                "Runway Bearing": int(round(bearing / 10.0) * 10 / 10)
                            })

    df_junction_bearing = pd.DataFrame(junction_bearing)
    return df_junction_bearing


def format_junction_bearing(path_net: str) -> pd.DataFrame:
    """ Format the junction bearing data from a SUMO network XML file into a matchup table."""

    # Generate the original junction bearing DataFrame
    junction_bearing = generate_junction_bearing(path_net)

    MatchupTable = copy.deepcopy(junction_bearing)
    MatchupTable.drop(columns=[
        "Second Last Point (x, y)", "Last Point (x, y)",
        "Second Last Point (lat, lon)", "Last Point (lat, lon)"
    ], inplace=True)
    MatchupTable.rename(columns={"Approach Edge": "FromEdge"}, inplace=True)
    MatchupTable.insert(MatchupTable.columns.get_loc("FromEdge") + 1, "ToEdge", None)
    MatchupTable.insert(MatchupTable.columns.get_loc("ToEdge") + 1, "Direction", None)

    df_connections = get_net_connections(path_net)
    MatchupTable = MatchupTable.merge(df_connections, on="FromEdge", how="left")
    MatchupTable.drop(columns=["ToEdge_x", "Direction_x"], inplace=True)
    MatchupTable.rename(columns={"ToEdge_y": "ToEdge", "Direction_y": "Direction"}, inplace=True)
    MatchupTable.drop_duplicates(subset=["Junction ID", "FromEdge", "ToEdge", "Direction"], inplace=True)
    MatchupTable.reset_index(drop=True, inplace=True)

    # MatchupTable["Junction ID Numeric"] = pd.to_numeric(MatchupTable["Junction ID"], errors='coerce')
    MatchupTable["Junction ID Numeric"] = MatchupTable["Junction ID"].astype(str)
    direction_order = {"right": 1, "thru": 2, "left": 3, "Uturn": 4}
    MatchupTable["Direction Order"] = MatchupTable["Direction"].map(direction_order)
    # Sort by "Junction ID", "Degree", and "Direction Order"
    MatchupTable = MatchupTable.sort_values(
        by=["Junction ID Numeric", "Degree", "Direction Order"],
        ascending=[True, True, True],
        na_position='last'
    )
    MatchupTable.drop(columns=["Junction ID Numeric", "Direction Order"], inplace=True)
    MatchupTable.reset_index(drop=True, inplace=True)

    MatchupTable.rename(columns={
        "Junction ID": "JunctionID_OpenDrive",
        "FromEdge": "FromRoadID_OpenDrive",
        "Degree": "Bearing",
        "Runway Bearing": "Numbering",
        "ToEdge": "ToRoadID_OpenDrive",
        "Direction": "Turn"
    }, inplace=True)

    # remove "-" before the sumo edge id to be OpenDrive road id
    MatchupTable["FromRoadID_OpenDrive"] = MatchupTable["FromRoadID_OpenDrive"].astype(str).str.lstrip("-")
    MatchupTable["ToRoadID_OpenDrive"] = MatchupTable["ToRoadID_OpenDrive"].astype(str).str.lstrip("-")

    MatchupTable["File_GridSmart"] = None
    MatchupTable["Date_GridSmart"] = None
    MatchupTable["IntersectionName_GridSmart"] = None
    MatchupTable["Turn_GridSmart"] = None
    MatchupTable["File_Synchro"] = None
    MatchupTable["IntersectionID_Synchro"] = None
    MatchupTable["Turn_Synchro"] = None
    MatchupTable["Need calibration?"] = None

    MatchupTable = MatchupTable[[
        "JunctionID_OpenDrive", "Bearing", "Numbering", "FromRoadID_OpenDrive",
        "ToRoadID_OpenDrive", "Turn", "File_GridSmart", "Date_GridSmart",
        "IntersectionName_GridSmart", "Turn_GridSmart", "File_Synchro",
        "IntersectionID_Synchro", "Turn_Synchro", "Need calibration?"
    ]]
    return MatchupTable


# Function to calculate bearing
def calculate_bearing(lat1, lon1, lat2, lon2):
    """ Calculate the bearing between two geographical points."""
    delta_lon = math.radians(lon2 - lon1)
    lat1 = math.radians(lat1)
    lat2 = math.radians(lat2)
    x = math.sin(delta_lon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(delta_lon)
    initial_bearing = math.atan2(x, y)
    initial_bearing = math.degrees(initial_bearing)
    compass_bearing = (initial_bearing + 360) % 360
    return compass_bearing


# Function to match input road to best bearing
def match_best_bearing(junction_id, lat1, lon1, lat2, lon2, junction_bearing):
    """ Match the input road to the best bearing based on the junction bearing DataFrame."""
    if not isinstance(junction_id, str):
        junction_id = str(junction_id)

    input_bearing = calculate_bearing(lat1, lon1, lat2, lon2)
    # runout_bearing = int(round(input_bearing / 10.0) * 10 / 10)

    # Filter junction bearings for the given junction
    matching_junctions = junction_bearing[junction_bearing["Junction ID"] == junction_id].copy()

    if matching_junctions.empty:
        print(f"No matching junction found for {junction_id}.")
        return

    # Find the best matching edge
    matching_junctions.loc[:, "Bearing Difference"] = abs(matching_junctions["Degree"] - input_bearing)
    best_match = matching_junctions.loc[matching_junctions["Bearing Difference"].idxmin()]

    # Output absolute difference instead of similarity
    abs_difference = best_match["Bearing Difference"]

    print(f"This road should be edge {best_match['Approach Edge']} "
          f"(runway bearing {int(round(best_match['Degree'] / 10.0) * 10 / 10)}) "
          f"of junction {junction_id}. The difference is {abs_difference:.2f}°.")

    return best_match['Approach Edge'], int(round(best_match['Degree'] / 10.0) * 10 / 10), abs_difference


# if __name__ == "__main__":
#     path_net = "./chatt.net.xml"
#     generate_matchup_table(path_net, "MatchupTable.xlsx")
